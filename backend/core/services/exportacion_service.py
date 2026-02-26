from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from core.models.equipamiento import Equipamiento
from extension import db
from core.models.personal import Personal, Becario, Investigador
from core.models.documentacion_autores import DocumentacionBibliografica
from core.models.grupo import GrupoInvestigacionUtn
from core.models.actividad_docencia import ActividadDocencia
from core.models.articulo_divulgacion import ArticuloDivulgacion
from core.models.becas import Beca
from core.models.directivos import DirectivoGrupo
from core.models.distinciones import DistincionRecibida
from core.models.erogacion import Erogacion
from core.models.participacion_relevante import ParticipacionRelevante
from core.models.registro_patente import RegistrosPropiedad
from  core.models.trabajo_reunion import TrabajoReunionCientifica
from core.models.trabajo_revista import TrabajosRevistasReferato
from core.models.transferencia_socio import TransferenciaSocioProductiva
from core.models.proyecto_investigacion import ProyectoInvestigacion

class ExportService:

    @staticmethod
    def generar_excel_grupo(grupo_id: int = 1):

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Grupo"

        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1

        # =====================================================
        # 1 - INDIVIDUALIZACIÓN
        # =====================================================
        grupo = GrupoInvestigacionUtn.query.get(grupo_id)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="1 - INDIVIDUALIZACIÓN DEL GRUPO").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        datos_grupo = [
            ["Unidad Académica", grupo.nombre_unidad_academica],
            ["Nombre y Sigla", grupo.nombre_sigla_grupo],
            ["Mail", grupo.mail],
            ["Objetivo", grupo.objetivo_desarrollo],
        ]

        for label, value in datos_grupo:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # =====================================================
        # 2 - PERSONAL
        # =====================================================
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="2 - PERSONAL").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 2

        # -----------------------------
        # 2.1 INVESTIGADORES
        # -----------------------------
        investigadores = Investigador.query.filter_by(
            grupo_utn_id=grupo_id,
            activo=True
        ).all()

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="2.1 - Investigadores").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        headers = ["N°", "Nombre y Apellido", "Categoría UTN", "Prog. Incentivos", "Dedicación", "Horas semanales"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin
        row += 1

        for i, inv in enumerate(investigadores, start=1):
            fila = [
                i,
                inv.nombre_apellido,
                inv.categoria_utn.nombre if inv.categoria_utn else "-",
                inv.programa_incentivos.nombre if inv.programa_incentivos else "-",
                inv.tipo_dedicacion.nombre if inv.tipo_dedicacion else "-",
                inv.horas_semanales
            ]
            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin
            row += 1

        row += 2

        # -----------------------------
        # 2.2 PERSONAL (separado por tipo)
        # -----------------------------
        personal = Personal.query.filter_by(
            grupo_utn_id=grupo_id,
            activo=True
        ).all()

        profesionales = [p for p in personal if p.tipo_personal.nombre.lower() == "profesional"]
        otros = [p for p in personal if p.tipo_personal.nombre.lower() != "profesional"]

        # Profesional
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="2.2 - Personal Profesional").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        headers_pp = ["N°", "Nombre y Apellido", "Horas semanales"]
        for col, h in enumerate(headers_pp, 1):
            ws.cell(row=row, column=col, value=h).font = bold
        row += 1

        for i, p in enumerate(profesionales, start=1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=p.nombre_apellido)
            ws.cell(row=row, column=3, value=p.horas_semanales)
            row += 1

        row += 2
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="2.3 - Personal técnico, administrativo y de apoyo").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1
        
        headers_pt = ["N°", "Nombre y Apellido", "Horas semanales"]
        for col, h in enumerate(headers_pt, 1):
            ws.cell(row=row, column=col, value=h).font = bold
        row += 1
        
        for i, p in enumerate(otros, start=1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=p.nombre_apellido)
            ws.cell(row=row, column=3, value=p.horas_semanales)
            row += 1
        row += 2

        # =====================================================
        # 2.4 - BECARIOS AGRUPADOS POR TIPO DE FORMACIÓN
        # =====================================================

        becarios = Becario.query.filter_by(
            grupo_utn_id=grupo_id,
            activo=True
        ).all()

        # Agrupar por tipo_formacion.nombre
        agrupados = {}
        for b in becarios:
            tipo = b.tipo_formacion.nombre if b.tipo_formacion else "Sin tipo"
            agrupados.setdefault(tipo, []).append(b)

        # Mapeo nombre BD → nombre plantilla
        mapa_titulos = {
            "Doctorado": "Doctorado",
            "Maestria": "Maestría / Especialización (EN CURSO)",
            "Graduado": "Becario Graduado",
            "Alumno": "Becarios Alumnos",
            "Pasante": "Pasantes",
            "Tesis de Posgrado": "Proyectos Finales y Tesis de Grado / Posgrado"
        }

        for tipo_bd, titulo_excel in mapa_titulos.items():

            lista = agrupados.get(tipo_bd, [])

            # Título sección
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1, value=titulo_excel).fill = yellow
            ws.cell(row=row, column=1).font = bold
            row += 1

            # Headers
            headers = ["N°", "Nombre y Apellido", "F. Financiamiento", "Horas semanales"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h).font = bold
            row += 1

            if not lista:
                ws.cell(row=row, column=1, value="No aplica.")
                row += 2
                continue

            for i, b in enumerate(lista, start=1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=b.nombre_apellido)
                ws.cell(row=row, column=3,
                        value=b.fuente_financiamiento.nombre if b.fuente_financiamiento else "Sin financiamiento")
                ws.cell(row=row, column=4, value=b.horas_semanales)
                row += 1

            row += 2

        # =====================================================
        # 4 - DOCUMENTACIÓN
        # =====================================================
        docs = DocumentacionBibliografica.query.filter_by(
            grupo_id=grupo_id
        ).all()

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="4 - DOCUMENTACIÓN Y BIBLIOTECA").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        headers_doc = ["N°", "Título", "Autores", "Editorial", "Año"]
        for col, h in enumerate(headers_doc, 1):
            ws.cell(row=row, column=col, value=h).font = bold
        row += 1

        for i, d in enumerate(docs, start=1):
            autores = ", ".join([a.nombre_apellido for a in d.autores])
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=d.titulo)
            ws.cell(row=row, column=3, value=autores)
            ws.cell(row=row, column=4, value=d.editorial)
            ws.cell(row=row, column=5, value=d.anio)
            row += 1

    
    
        # =====================================================
        # ACTIVIDADES DE DOCENCIA
        # =====================================================
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="5 - ACTIVIDADES DE DOCENCIA").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        actividades = (
            db.session.query(ActividadDocencia)
            .join(Investigador)
            .filter(Investigador.grupo_utn_id == grupo_id)
            .all()
        )

        headers = ["N°", "Curso", "Institución", "Investigador", "Grado Académico", "Rol", "Inicio", "Fin"]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, act in enumerate(actividades, start=1):
            fila = [
                i,
                act.curso,
                act.institucion,
                act.investigador.nombre_apellido if act.investigador else "-",
                act.grado_academico.nombre if act.grado_academico else "-",
                act.rol_actividad.nombre if act.rol_actividad else "-",
                act.fecha_inicio,
                act.fecha_fin
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin
            row += 1

        row += 2
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="6 - ARTÍCULOS DE DIVULGACIÓN").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        articulos = ArticuloDivulgacion.query.filter_by(grupo_utn_id=grupo_id).all()

        headers = ["N°", "Título", "Descripción", "Fecha Publicación"]

        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = bold

        row += 1

        for i, art in enumerate(articulos, start=1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=art.titulo)
            ws.cell(row=row, column=3, value=art.descripcion)
            ws.cell(row=row, column=4, value=art.fecha_publicacion)
            row += 1

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="7 - BECAS Y BECARIOS").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        becas = Beca.query.all()

        headers = ["N°", "Beca", "Fuente Financiamiento", "Becario", "Inicio", "Fin", "Monto"]

        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = bold

        row += 1

        contador = 1

        for beca in becas:
            for relacion in beca.becarios:
                if relacion.becario.grupo_utn_id != grupo_id:
                    continue

                fila = [
                    contador,
                    beca.nombre_beca,
                    beca.fuente_financiamiento.nombre if beca.fuente_financiamiento else "-",
                    relacion.becario.nombre_apellido,
                    relacion.fecha_inicio,
                    relacion.fecha_fin,
                    relacion.monto_percibido
                ]

                for col, val in enumerate(fila, 1):
                    ws.cell(row=row, column=col, value=val)

                row += 1
                contador += 1

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="8 - DIRECTIVOS").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        directivos = DirectivoGrupo.query.filter_by(id_grupo_utn=grupo_id).all()

        headers = ["N°", "Nombre", "Cargo", "Inicio", "Fin"]

        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = bold

        row += 1

        for i, d in enumerate(directivos, start=1):
            fila = [
                i,
                d.directivo.nombre_apellido,
                d.cargo.nombre,
                d.fecha_inicio,
                d.fecha_fin
            ]

            for col, val in enumerate(fila, 1):
                ws.cell(row=row, column=col, value=val)

            row += 1

        row += 2


        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value="9 - DISTINCIONES RECIBIDAS").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        distinciones = DistincionRecibida.query.all()

        headers = ["N°", "Fecha", "Descripción", "Proyecto"]

        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = bold

        row += 1

        contador = 1

        for d in distinciones:
            if not d.proyecto_investigacion:
                continue

            if d.proyecto_investigacion.grupo_utn_id != grupo_id:
                continue

            fila = [
                contador,
                d.fecha,
                d.descripcion,
                d.proyecto_investigacion.nombre_proyecto
            ]

            for col, val in enumerate(fila, 1):
                ws.cell(row=row, column=col, value=val)

            row += 1
            contador += 1

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="10 - EQUIPAMIENTO DEL GRUPO").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        equipamientos = Equipamiento.query.filter_by(
            grupo_utn_id=grupo_id
        ).all()

        headers = [
            "N°",
            "Denominación",
            "Descripción",
            "Fecha Incorporación",
            "Monto Invertido"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, eq in enumerate(equipamientos, start=1):
            fila = [
                i,
                eq.denominacion,
                eq.descripcion_breve,
                eq.fecha_incorporacion,
                eq.monto_invertido
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            row += 1

        row += 2


        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value="11 - EROGACIONES DEL GRUPO").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        erogaciones = Erogacion.query.filter(
            Erogacion.grupo_utn_id == grupo_id,
            Erogacion.deleted_at.is_(None)
        ).all()

        headers = [
            "N°",
            "N° Erogación",
            "Fecha",
            "Tipo",
            "Fuente Financiamiento",
            "Ingresos",
            "Egresos",
            "Saldo"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        total_ingresos = 0
        total_egresos = 0

        for i, e in enumerate(erogaciones, start=1):

            saldo = e.ingresos - e.egresos

            fila = [
                i,
                e.numero_erogacion,
                e.fecha,
                e.tipo_erogacion.nombre if e.tipo_erogacion else "-",
                e.fuente_financiamiento.nombre if e.fuente_financiamiento else "-",
                e.ingresos,
                e.egresos,
                saldo
            ]

            total_ingresos += e.ingresos
            total_egresos += e.egresos

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            row += 1

        row += 1

        # -------------------------
        # TOTALES
        # -------------------------

        ws.cell(row=row, column=5, value="TOTALES").font = bold

        ws.cell(row=row, column=6, value=total_ingresos).font = bold
        ws.cell(row=row, column=7, value=total_egresos).font = bold
        ws.cell(row=row, column=8, value=(total_ingresos - total_egresos)).font = bold

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="12 - PARTICIPACIONES RELEVANTES").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        participaciones = (
            db.session.query(ParticipacionRelevante)
            .join(Investigador)
            .filter(
                Investigador.grupo_utn_id == grupo_id,
                Investigador.deleted_at.is_(None)
            )
            .order_by(ParticipacionRelevante.fecha.desc())
            .all()
        )

        headers = [
            "N°",
            "Evento",
            "Forma de Participación",
            "Fecha",
            "Investigador"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, p in enumerate(participaciones, start=1):
            fila = [
                i,
                p.nombre_evento,
                p.forma_participacion,
                p.fecha,
                p.investigador.nombre_apellido if p.investigador else "-"
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            row += 1

        row += 2
        
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="13 - REGISTROS DE PROPIEDAD INTELECTUAL").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        registros = RegistrosPropiedad.query.filter(
            RegistrosPropiedad.grupo_utn_id == grupo_id,
            RegistrosPropiedad.deleted_at.is_(None)
        ).order_by(RegistrosPropiedad.fecha_registro.desc()).all()

        headers = [
            "N°",
            "Nombre / Título",
            "Tipo de Registro",
            "Organismo Registrante",
            "Fecha de Registro"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, r in enumerate(registros, start=1):
            fila = [
                i,
                r.nombre_articulo,
                r.tipo_registro.nombre if r.tipo_registro else "-",
                r.organismo_registrante,
                r.fecha_registro
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            # Formato fecha argentino opcional
            ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"

            row += 1

        row += 2


        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value="14 - TRABAJOS EN REUNIONES CIENTÍFICAS").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        trabajos = TrabajoReunionCientifica.query.filter(
            TrabajoReunionCientifica.grupo_utn_id == grupo_id,
            TrabajoReunionCientifica.deleted_at.is_(None)
        ).order_by(TrabajoReunionCientifica.fecha_inicio.desc()).all()

        headers = [
            "N°",
            "Título del Trabajo",
            "Reunión Científica",
            "Tipo",
            "Procedencia",
            "Fecha",
            "Investigadores"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, t in enumerate(trabajos, start=1):

            investigadores_str = ", ".join(
                [inv.nombre_apellido for inv in t.investigadores if inv.deleted_at is None]
            )

            fila = [
                i,
                t.titulo_trabajo,
                t.nombre_reunion,
                t.tipo_reunion_cientifica.nombre if t.tipo_reunion_cientifica else "-",
                t.procedencia,
                t.fecha_inicio,
                investigadores_str if investigadores_str else "-"
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            # formato fecha
            ws.cell(row=row, column=6).number_format = "DD/MM/YYYY"

            row += 1

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row=row, column=1, value="15 - TRABAJOS EN REVISTAS CON REFERATO").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        trabajos_revistas = TrabajosRevistasReferato.query.filter(
            TrabajosRevistasReferato.grupo_utn_id == grupo_id,
            TrabajosRevistasReferato.deleted_at.is_(None)
        ).order_by(TrabajosRevistasReferato.fecha.desc()).all()

        headers = [
            "N°",
            "Título",
            "Revista",
            "Editorial",
            "ISSN",
            "País",
            "Tipo",
            "Fecha",
            "Investigadores"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        for i, t in enumerate(trabajos_revistas, start=1):

            investigadores_str = ", ".join(
                [inv.nombre_apellido for inv in t.investigadores if inv.deleted_at is None]
            )

            fila = [
                i,
                t.titulo_trabajo,
                t.nombre_revista,
                t.editorial,
                t.issn,
                t.pais,
                t.tipo_reunion.nombre if t.tipo_reunion else "-",
                t.fecha,
                investigadores_str if investigadores_str else "-"
            ]

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            # Formato fecha
            ws.cell(row=row, column=8).number_format = "DD/MM/YYYY"

            row += 1

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row=row, column=1, value="16 - TRANSFERENCIAS SOCIO-PRODUCTIVAS").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        transferencias = TransferenciaSocioProductiva.query.filter(
            TransferenciaSocioProductiva.grupo_utn_id == grupo_id,
            TransferenciaSocioProductiva.deleted_at.is_(None)
        ).order_by(TransferenciaSocioProductiva.fecha_inicio.desc()).all()

        headers = [
            "N°",
            "N° Transferencia",
            "Denominación",
            "Demandante",
            "Tipo Contrato",
            "Monto",
            "Fecha Inicio",
            "Fecha Fin",
            "Adoptantes"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        total_monto = 0

        for i, t in enumerate(transferencias, start=1):

            adoptantes_str = ", ".join(
                [
                    p.adoptante.nombre
                    for p in t.participaciones
                    if p.deleted_at is None
                ]
            )

            fila = [
                i,
                t.numero_transferencia,
                t.denominacion,
                t.demandante,
                t.tipo_contrato_transferencia.nombre if t.tipo_contrato_transferencia else "-",
                t.monto,
                t.fecha_inicio,
                t.fecha_fin,
                adoptantes_str if adoptantes_str else "-"
            ]

            if t.monto:
                total_monto += t.monto

            for col, val in enumerate(fila, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = thin

            # Formato fechas
            ws.cell(row=row, column=7).number_format = "DD/MM/YYYY"
            ws.cell(row=row, column=8).number_format = "DD/MM/YYYY"

            # Formato moneda
            ws.cell(row=row, column=6).number_format = '"$"#,##0.00'

            row += 1

        row += 1

        # -------------------------
        # TOTAL MONTO TRANSFERENCIAS
        # -------------------------

        ws.cell(row=row, column=4, value="TOTAL MONTO TRANSFERENCIAS").font = bold
        ws.cell(row=row, column=6, value=total_monto).font = bold
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'

        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.cell(row=row, column=1, value="17 - PROYECTOS DE INVESTIGACIÓN").fill = yellow
        ws.cell(row=row, column=1).font = bold
        row += 1

        proyectos = ProyectoInvestigacion.query.filter(
            ProyectoInvestigacion.grupo_utn_id == grupo_id,
            ProyectoInvestigacion.deleted_at.is_(None)
        ).order_by(ProyectoInvestigacion.fecha_inicio.desc()).all()

        headers = [
            "N°",
            "Código",
            "Nombre",
            "Tipo",
            "Fuente",
            "Monto Destinado",
            "Fecha Inicio",
            "Fecha Fin",
            "Investigadores",
            "Becarios"
        ]

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold
            c.border = thin

        row += 1

        total_monto_proyectos = 0

        for i, p in enumerate(proyectos, start=1):

            # -------------------------
            # FILA PRINCIPAL PROYECTO
            # -------------------------
            ws.cell(row=row, column=1, value=i).border = thin
            ws.cell(row=row, column=2, value=p.codigo_proyecto).border = thin
            ws.cell(row=row, column=3, value=p.nombre_proyecto).border = thin
            ws.cell(row=row, column=4, value=p.tipo_proyecto.nombre if p.tipo_proyecto else "-").border = thin
            ws.cell(row=row, column=5, value=p.fuente_financiamiento.nombre if p.fuente_financiamiento else "-").border = thin
            ws.cell(row=row, column=6, value=p.monto_destinado).border = thin
            ws.cell(row=row, column=7, value=p.fecha_inicio).border = thin
            ws.cell(row=row, column=8, value=p.fecha_fin).border = thin

            ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=7).number_format = "DD/MM/YYYY"
            ws.cell(row=row, column=8).number_format = "DD/MM/YYYY"

            row += 1

            # -------------------------
            # INVESTIGADORES
            # -------------------------
            ws.cell(row=row, column=2, value="Investigadores:").font = bold
            row += 1

            for rel in p.participaciones_investigador:
                if rel.deleted_at is None and rel.investigador:
                    ws.cell(row=row, column=3, value=rel.investigador.nombre_apellido)
                    ws.cell(row=row, column=4, value=rel.fecha_inicio)
                    ws.cell(row=row, column=5, value=rel.fecha_fin)
                    ws.cell(row=row, column=4).number_format = "DD/MM/YYYY"
                    ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"
                    row += 1

            # -------------------------
            # BECARIOS
            # -------------------------
            ws.cell(row=row, column=2, value="Becarios:").font = bold
            row += 1

            for rel in p.participaciones_becario:
                if rel.deleted_at is None and rel.becario:
                    ws.cell(row=row, column=3, value=rel.becario.nombre_apellido)
                    ws.cell(row=row, column=4, value=rel.fecha_inicio)
                    ws.cell(row=row, column=5, value=rel.fecha_fin)
                    ws.cell(row=row, column=4).number_format = "DD/MM/YYYY"
                    ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"
                    row += 1

            row += 2

        # -------------------------
        # TOTAL MONTO PROYECTOS
        # -------------------------

        ws.cell(row=row, column=4, value="TOTAL MONTO PROYECTOS").font = bold
        ws.cell(row=row, column=6, value=total_monto_proyectos).font = bold
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'

        row += 2
            
        
            # Guardar en memoria        output = BytesIO()
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    
        return output
    