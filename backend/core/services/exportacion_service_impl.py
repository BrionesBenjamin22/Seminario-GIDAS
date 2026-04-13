from io import BytesIO
from datetime import date
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import joinedload, selectinload

from core.models.actividad_docencia import ActividadDocencia, InvestigadorActividadGrado
from core.models.articulo_divulgacion import ArticuloDivulgacion
from core.models.becas import Beca, Beca_Becario
from core.models.directivos import DirectivoGrupo
from core.models.distinciones import DistincionRecibida
from core.models.documentacion_autores import DocumentacionBibliografica
from core.models.equipamiento import Equipamiento
from core.models.erogacion import Erogacion
from core.models.grupo import GrupoInvestigacionUtn
from core.models.participacion_relevante import ParticipacionRelevante
from core.models.personal import Becario, Investigador, Personal
from core.models.proyecto_investigacion import ProyectoInvestigacion, InvestigadorProyecto, BecarioProyecto
from core.models.registro_patente import RegistrosPropiedad
from core.models.trabajo_reunion import TrabajoReunionCientifica
from core.models.trabajo_revista import TrabajosRevistasReferato
from core.models.transferencia_socio import TransferenciaSocioProductiva, AdoptanteTransferencia
from core.models.visita_grupo import VisitaAcademica
from core.models.programa_actividades import PlanificacionGrupo


class ExportService:
    TITLE_FILL = PatternFill(fill_type="solid", fgColor="FBE4D5")
    SECTION_FILL = PatternFill(fill_type="solid", fgColor="FBE4D5")
    SUBSECTION_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
    ACCENT_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
    TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
    THIN_SIDE = Side(style="thin", color="000000")
    THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    HEADER_FONT = Font(name="Calibri", size=11, bold=True)
    TITLE_FONT = Font(name="Calibri", size=15, bold=True)
    BODY_FONT = Font(name="Calibri", size=11)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    COLUMN_WIDTHS = {
        "A": 23.14, "B": 40.57, "C": 36.43, "D": 23.14,
        "E": 27.43, "F": 34.14, "G": 65.14, "H": 74.86,
        "I": 81.43, "J": 29.0, "K": 13.0, "L": 13.0,
    }

    @staticmethod
    def _current_hours(entity):
        historial = getattr(entity, "historial_horas", None) or []
        activo = next((h for h in historial if h.fecha_fin is None), None)
        return activo.horas_semanales if activo else getattr(entity, "horas_semanales", None)

    @staticmethod
    def _active_grado_nombre(actividad: ActividadDocencia):
        historial = getattr(actividad, "investigadores_grado", None) or []
        grado_activo = next((h.grado_academico for h in historial if h.fecha_fin is None), None)
        return grado_activo.nombre if grado_activo else "-"

    @staticmethod
    def _money(value):
        return float(value or 0)

    @staticmethod
    def _join_names(items: Iterable, attr: str = "nombre_apellido", fallback: str = "-"):
        values = []
        for item in items:
            if getattr(item, "deleted_at", None) is None:
                value = getattr(item, attr, None)
                if value:
                    values.append(str(value))
        return ", ".join(values) if values else fallback

    @staticmethod
    def _safe_text(value, fallback="-"):
        if value is None:
            return fallback
        if isinstance(value, str):
            value = " ".join(value.split())
            return value if value else fallback
        return value

    @classmethod
    def _build_workbook(cls):
        wb = Workbook()
        ws = wb.active
        ws.title = "Memorias"
        ws.freeze_panes = "A4"
        for column, width in cls.COLUMN_WIDTHS.items():
            ws.column_dimensions[column].width = width
        return wb, ws

    @classmethod
    def _style_cell(cls, cell, *, font=None, fill=None, border=None, alignment=None, number_format=None):
        cell.font = font or cls.BODY_FONT
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format
        return cell

    @classmethod
    def _write_title(cls, ws, row: int, text: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.TITLE_FONT,
            fill=cls.TITLE_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.CENTER,
        )
        ws.row_dimensions[row].height = 24
        return row + 2

    @classmethod
    def _write_section(cls, ws, row: int, text: str, span: int = 12):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.HEADER_FONT,
            fill=cls.SECTION_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        return row + 2

    @classmethod
    def _write_subsection(cls, ws, row: int, text: str, span: int = 8, accent: bool = False):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws.cell(row=row, column=1, value=text)
        cls._style_cell(
            cell,
            font=cls.HEADER_FONT,
            fill=cls.ACCENT_FILL if accent else cls.SUBSECTION_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        return row + 2

    @classmethod
    def _write_label_value(cls, ws, row: int, label: str, value, value_span: int = 5):
        label_cell = ws.cell(row=row, column=1, value=label)
        cls._style_cell(label_cell, font=cls.HEADER_FONT, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + value_span)
        value_cell = ws.cell(row=row, column=2, value=value)
        cls._style_cell(value_cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        return row + 1

    @classmethod
    def _write_multiline_block(cls, ws, row: int, label: str, text: str, height: int = 72):
        row = cls._write_subsection(ws, row, label, span=8)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=8)
        cell = ws.cell(row=row, column=1, value=cls._safe_text(text, fallback="Sin informacion registrada."))
        cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
        ws.row_dimensions[row].height = height
        return row + 5

    @classmethod
    def _write_table(cls, ws, row: int, title: str, headers: Sequence[str], rows: Sequence[Sequence], *, accent: bool = False, merge_span: int | None = None, date_cols: set[int] | None = None, money_cols: set[int] | None = None):
        row = cls._write_subsection(ws, row, title, span=merge_span or max(len(headers), 4), accent=accent)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cls._style_cell(cell, font=cls.HEADER_FONT, border=cls.THIN_BORDER, alignment=cls.CENTER)
        ws.row_dimensions[row].height = 28
        row += 1
        if not rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(len(headers), 4))
            cell = ws.cell(row=row, column=1, value="No registra datos.")
            cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP)
            return row + 3
        date_cols = date_cols or set()
        money_cols = money_cols or set()
        for row_data in rows:
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                number_format = None
                if col_idx in date_cols and value is not None:
                    number_format = "DD/MM/YYYY"
                elif col_idx in money_cols and value is not None:
                    number_format = '"$"#,##0.00'
                cls._style_cell(cell, border=cls.THIN_BORDER, alignment=cls.LEFT_TOP if col_idx != 1 else cls.CENTER, number_format=number_format)
            row += 1
        return row + 2

    @classmethod
    def _write_totals(cls, ws, row: int, label: str, values: Sequence[tuple[int, float]]):
        label_cell = ws.cell(row=row, column=1, value=label)
        cls._style_cell(
            label_cell,
            font=cls.HEADER_FONT,
            fill=cls.TOTAL_FILL,
            border=cls.THIN_BORDER,
            alignment=cls.LEFT_TOP,
        )
        for offset, (_, value) in enumerate(values, start=2):
            cell = ws.cell(row=row, column=offset, value=value)
            cls._style_cell(
                cell,
                font=cls.HEADER_FONT,
                fill=cls.TOTAL_FILL,
                border=cls.THIN_BORDER,
                alignment=cls.CENTER,
                number_format='"$"#,##0.00'
            )
        return row + 3

    @classmethod
    def _write_grouped_tables(
        cls,
        ws,
        row: int,
        base_index: str,
        base_title: str,
        groups: Sequence[tuple[str, Sequence[Sequence]]],
        headers: Sequence[str],
        *,
        accent: bool = False,
        merge_span: int | None = None,
        date_cols: set[int] | None = None,
        money_cols: set[int] | None = None,
    ):
        row = cls._write_subsection(
            ws,
            row,
            f"{base_index}.- {base_title}",
            span=merge_span or max(len(headers), 4),
            accent=accent,
        )

        subgroup_number = 1
        for group_title, group_rows in groups:
            if not group_rows:
                continue
            row = cls._write_table(
                ws,
                row,
                f"{base_index}.{subgroup_number}.- {group_title}",
                headers,
                group_rows,
                accent=accent,
                merge_span=merge_span,
                date_cols=date_cols,
                money_cols=money_cols,
            )
            subgroup_number += 1

        if subgroup_number == 1:
            row = cls._write_table(
                ws,
                row,
                f"{base_index}.1.- Sin registros",
                headers,
                [],
                accent=accent,
                merge_span=merge_span,
                date_cols=date_cols,
                money_cols=money_cols,
            )

        return row

    @staticmethod
    def _get_grupo(grupo_id: int | None):
        query = GrupoInvestigacionUtn.query.filter(GrupoInvestigacionUtn.deleted_at.is_(None))
        grupo = query.filter(GrupoInvestigacionUtn.id == grupo_id).first() if grupo_id is not None else query.order_by(GrupoInvestigacionUtn.id.asc()).first()
        if not grupo:
            raise ValueError("Grupo UTN no encontrado")
        return grupo

    @staticmethod
    def _get_directivos(grupo_id: int):
        return (
            DirectivoGrupo.query.options(joinedload(DirectivoGrupo.directivo), joinedload(DirectivoGrupo.cargo))
            .filter(DirectivoGrupo.id_grupo_utn == grupo_id, DirectivoGrupo.deleted_at.is_(None))
            .order_by(DirectivoGrupo.fecha_inicio.asc(), DirectivoGrupo.id.asc())
            .all()
        )

    @staticmethod
    def _get_investigadores(grupo_id: int):
        return (
            Investigador.query.options(joinedload(Investigador.categoria_utn), joinedload(Investigador.programa_incentivos), joinedload(Investigador.tipo_dedicacion), selectinload(Investigador.historial_horas))
            .filter(Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(Investigador.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_personal(grupo_id: int):
        return (
            Personal.query.options(joinedload(Personal.tipo_personal), selectinload(Personal.historial_horas))
            .filter(Personal.grupo_utn_id == grupo_id, Personal.deleted_at.is_(None))
            .order_by(Personal.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_becarios(grupo_id: int):
        return (
            Becario.query.options(joinedload(Becario.tipo_formacion), selectinload(Becario.historial_horas), selectinload(Becario.becas).joinedload(Beca_Becario.beca).joinedload(Beca.fuente_financiamiento))
            .filter(Becario.grupo_utn_id == grupo_id, Becario.deleted_at.is_(None))
            .order_by(Becario.nombre_apellido.asc())
            .all()
        )

    @staticmethod
    def _get_documentacion(grupo_id: int):
        return (
            DocumentacionBibliografica.query.options(selectinload(DocumentacionBibliografica.autores))
            .filter(DocumentacionBibliografica.grupo_id == grupo_id, DocumentacionBibliografica.deleted_at.is_(None))
            .order_by(DocumentacionBibliografica.anio.desc(), DocumentacionBibliografica.titulo.asc())
            .all()
        )

    @staticmethod
    def _get_actividades_docencia(grupo_id: int):
        return (
            ActividadDocencia.query.options(joinedload(ActividadDocencia.investigador), joinedload(ActividadDocencia.rol_actividad), selectinload(ActividadDocencia.investigadores_grado).joinedload(InvestigadorActividadGrado.grado_academico))
            .join(Investigador, ActividadDocencia.investigador_id == Investigador.id)
            .filter(ActividadDocencia.deleted_at.is_(None), Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(ActividadDocencia.fecha_inicio.desc(), ActividadDocencia.id.desc())
            .all()
        )

    @staticmethod
    def _get_articulos(grupo_id: int):
        return (
            ArticuloDivulgacion.query.filter(ArticuloDivulgacion.grupo_utn_id == grupo_id, ArticuloDivulgacion.deleted_at.is_(None))
            .order_by(ArticuloDivulgacion.fecha_publicacion.desc(), ArticuloDivulgacion.id.desc())
            .all()
        )

    @staticmethod
    def _get_becas(grupo_id: int):
        becas = (
            Beca.query.options(joinedload(Beca.fuente_financiamiento), selectinload(Beca.becarios).joinedload(Beca_Becario.becario))
            .filter(Beca.deleted_at.is_(None))
            .order_by(Beca.nombre_beca.asc())
            .all()
        )
        rows = []
        for beca in becas:
            for relacion in beca.becarios:
                becario = relacion.becario
                if relacion.deleted_at is None and becario is not None and becario.deleted_at is None and becario.grupo_utn_id == grupo_id:
                    rows.append((beca, relacion))
        return rows

    @staticmethod
    def _get_distinciones(grupo_id: int):
        return (
            DistincionRecibida.query.options(joinedload(DistincionRecibida.proyecto_investigacion))
            .join(ProyectoInvestigacion, DistincionRecibida.proyecto_investigacion_id == ProyectoInvestigacion.id)
            .filter(DistincionRecibida.deleted_at.is_(None), ProyectoInvestigacion.grupo_utn_id == grupo_id, ProyectoInvestigacion.deleted_at.is_(None))
            .order_by(DistincionRecibida.fecha.desc(), DistincionRecibida.id.desc())
            .all()
        )

    @staticmethod
    def _get_equipamiento(grupo_id: int):
        return (
            Equipamiento.query.filter(Equipamiento.grupo_utn_id == grupo_id, Equipamiento.deleted_at.is_(None))
            .order_by(Equipamiento.fecha_incorporacion.desc(), Equipamiento.id.desc())
            .all()
        )

    @staticmethod
    def _get_erogaciones(grupo_id: int):
        return (
            Erogacion.query.options(joinedload(Erogacion.tipo_erogacion), joinedload(Erogacion.fuente_financiamiento))
            .filter(Erogacion.grupo_utn_id == grupo_id, Erogacion.deleted_at.is_(None))
            .order_by(Erogacion.fecha.desc(), Erogacion.id.desc())
            .all()
        )

    @staticmethod
    def _get_participaciones(grupo_id: int):
        return (
            ParticipacionRelevante.query.options(joinedload(ParticipacionRelevante.investigador))
            .join(Investigador, ParticipacionRelevante.investigador_id == Investigador.id)
            .filter(ParticipacionRelevante.deleted_at.is_(None), Investigador.grupo_utn_id == grupo_id, Investigador.deleted_at.is_(None))
            .order_by(ParticipacionRelevante.fecha.desc(), ParticipacionRelevante.id.desc())
            .all()
        )

    @staticmethod
    def _get_registros(grupo_id: int):
        return (
            RegistrosPropiedad.query.options(joinedload(RegistrosPropiedad.tipo_registro))
            .filter(RegistrosPropiedad.grupo_utn_id == grupo_id, RegistrosPropiedad.deleted_at.is_(None))
            .order_by(RegistrosPropiedad.fecha_registro.desc(), RegistrosPropiedad.id.desc())
            .all()
        )

    @staticmethod
    def _get_trabajos_reunion(grupo_id: int):
        return (
            TrabajoReunionCientifica.query.options(joinedload(TrabajoReunionCientifica.tipo_reunion_cientifica), selectinload(TrabajoReunionCientifica.investigadores))
            .filter(TrabajoReunionCientifica.grupo_utn_id == grupo_id, TrabajoReunionCientifica.deleted_at.is_(None))
            .order_by(TrabajoReunionCientifica.fecha_inicio.desc(), TrabajoReunionCientifica.id.desc())
            .all()
        )

    @staticmethod
    def _get_trabajos_revista(grupo_id: int):
        return (
            TrabajosRevistasReferato.query.options(joinedload(TrabajosRevistasReferato.tipo_reunion), selectinload(TrabajosRevistasReferato.investigadores))
            .filter(TrabajosRevistasReferato.grupo_utn_id == grupo_id, TrabajosRevistasReferato.deleted_at.is_(None))
            .order_by(TrabajosRevistasReferato.fecha.desc(), TrabajosRevistasReferato.id.desc())
            .all()
        )

    @staticmethod
    def _get_transferencias(grupo_id: int):
        return (
            TransferenciaSocioProductiva.query.options(joinedload(TransferenciaSocioProductiva.tipo_contrato_transferencia), selectinload(TransferenciaSocioProductiva.participaciones).joinedload(AdoptanteTransferencia.adoptante))
            .filter(TransferenciaSocioProductiva.grupo_utn_id == grupo_id, TransferenciaSocioProductiva.deleted_at.is_(None))
            .order_by(TransferenciaSocioProductiva.fecha_inicio.desc(), TransferenciaSocioProductiva.id.desc())
            .all()
        )

    @staticmethod
    def _get_visitas(grupo_id: int):
        return (
            VisitaAcademica.query.options(joinedload(VisitaAcademica.tipo_visita))
            .filter(
                VisitaAcademica.grupo_utn_id == grupo_id,
                VisitaAcademica.deleted_at.is_(None),
            )
            .order_by(VisitaAcademica.fecha.desc(), VisitaAcademica.id.desc())
            .all()
        )

    @staticmethod
    def _get_planificaciones(grupo_id: int):
        return (
            PlanificacionGrupo.query.filter(
                PlanificacionGrupo.grupo_id == grupo_id,
                PlanificacionGrupo.deleted_at.is_(None),
            )
            .order_by(PlanificacionGrupo.anio.asc(), PlanificacionGrupo.id.asc())
            .all()
        )

    @staticmethod
    def _get_proyectos(grupo_id: int):
        return (
            ProyectoInvestigacion.query.options(joinedload(ProyectoInvestigacion.tipo_proyecto), joinedload(ProyectoInvestigacion.fuente_financiamiento), selectinload(ProyectoInvestigacion.participaciones_investigador).joinedload(InvestigadorProyecto.investigador), selectinload(ProyectoInvestigacion.participaciones_becario).joinedload(BecarioProyecto.becario), selectinload(ProyectoInvestigacion.distinciones))
            .filter(ProyectoInvestigacion.grupo_utn_id == grupo_id, ProyectoInvestigacion.deleted_at.is_(None))
            .order_by(ProyectoInvestigacion.fecha_inicio.desc(), ProyectoInvestigacion.id.desc())
            .all()
        )

    @staticmethod
    def _find_directivo_nombre(directivos, fragments: Sequence[str]):
        for participacion in directivos:
            if participacion.deleted_at is not None:
                continue
            cargo = participacion.cargo.nombre.lower() if participacion.cargo and participacion.cargo.nombre else ""
            if any(fragment in cargo for fragment in fragments):
                if participacion.directivo and participacion.directivo.deleted_at is None:
                    return participacion.directivo.nombre_apellido
        return "-"

    @classmethod
    def generar_excel_grupo(cls, grupo_id: int | None = 1):
        grupo = cls._get_grupo(grupo_id)
        wb, ws = cls._build_workbook()

        directivos = cls._get_directivos(grupo.id)
        investigadores = cls._get_investigadores(grupo.id)
        personal = cls._get_personal(grupo.id)
        becarios = cls._get_becarios(grupo.id)
        documentacion = cls._get_documentacion(grupo.id)
        actividades = cls._get_actividades_docencia(grupo.id)
        articulos = cls._get_articulos(grupo.id)
        becas = cls._get_becas(grupo.id)
        distinciones = cls._get_distinciones(grupo.id)
        equipamiento = cls._get_equipamiento(grupo.id)
        erogaciones = cls._get_erogaciones(grupo.id)
        participaciones = cls._get_participaciones(grupo.id)
        registros = cls._get_registros(grupo.id)
        trabajos_reunion = cls._get_trabajos_reunion(grupo.id)
        trabajos_revista = cls._get_trabajos_revista(grupo.id)
        transferencias = cls._get_transferencias(grupo.id)
        visitas = cls._get_visitas(grupo.id)
        planificaciones = cls._get_planificaciones(grupo.id)
        proyectos = cls._get_proyectos(grupo.id)

        row = 1
        row = cls._write_title(ws, row, f"MEMORIAS {date.today().year} DEL GRUPO UTN - {grupo.nombre_sigla_grupo}")
        row = cls._write_section(ws, row, "I.- ADMINISTRACION")
        row = cls._write_subsection(ws, row, "1.- INDIVIDUALIZACION DEL GRUPO UTN", span=8)
        row = cls._write_label_value(ws, row, "1.1.- Facultad Regional", grupo.nombre_unidad_academica)
        row = cls._write_label_value(ws, row, "1.2.- Nombre y Sigla", grupo.nombre_sigla_grupo)
        row = cls._write_label_value(ws, row, "1.3.- Director/a", cls._find_directivo_nombre(directivos, ["director"]))
        row = cls._write_label_value(ws, row, "1.4.- Vicedirector/a", cls._find_directivo_nombre(directivos, ["vicedirector", "vice director"]))
        row = cls._write_label_value(ws, row, "1.5.- Direccion de Email", grupo.mail)
        row += 1
        directivos_rows = [[idx, p.directivo.nombre_apellido if p.directivo and p.directivo.deleted_at is None else "-", p.cargo.nombre if p.cargo else "-", p.fecha_inicio, p.fecha_fin] for idx, p in enumerate(directivos, start=1) if p.directivo and p.directivo.deleted_at is None]
        row = cls._write_table(ws, row, "1.6.- Autoridades y cargos de gestion del grupo", ["Nro.", "Apellido y nombre", "Cargo desempenado", "Fecha de inicio", "Fecha de finalizacion"], directivos_rows, merge_span=8, date_cols={4, 5})
        row = cls._write_multiline_block(ws, row, "1.7.- Objetivos y desarrollo", grupo.objetivo_desarrollo)
        row = cls._write_subsection(ws, row, "2.- PERSONAL", span=10)
        investigadores_rows = [[idx, inv.nombre_apellido, inv.categoria_utn.nombre if inv.categoria_utn else "-", inv.programa_incentivos.nombre if inv.programa_incentivos else "-", inv.tipo_dedicacion.nombre if inv.tipo_dedicacion else "-", cls._current_hours(inv)] for idx, inv in enumerate(investigadores, start=1)]
        row = cls._write_table(ws, row, "2.1.- Investigadores", ["Nro.", "Apellido y nombre", "Categoria UTN", "Programa de incentivos", "Tipo de dedicacion", "Carga horaria semanal"], investigadores_rows, merge_span=8)
        personal_rows = [[idx, persona.nombre_apellido, persona.tipo_personal.nombre if persona.tipo_personal else "-", cls._current_hours(persona)] for idx, persona in enumerate(personal, start=1)]
        row = cls._write_table(ws, row, "2.2.- Personal tecnico, administrativo y de apoyo", ["Nro.", "Apellido y nombre", "Categoria de personal", "Carga horaria semanal"], personal_rows, merge_span=8)

        becarios_rows = []
        for idx, becario in enumerate(becarios, start=1):
            relaciones_activas = [rel for rel in becario.becas if rel.deleted_at is None and rel.beca and rel.beca.deleted_at is None]
            nombres_becas = ", ".join(sorted({rel.beca.nombre_beca for rel in relaciones_activas})) or "-"
            fuentes = ", ".join(sorted({rel.beca.fuente_financiamiento.nombre for rel in relaciones_activas if rel.beca.fuente_financiamiento is not None})) or "-"
            becarios_rows.append([idx, becario.nombre_apellido, becario.tipo_formacion.nombre if becario.tipo_formacion else "-", nombres_becas, fuentes, cls._current_hours(becario)])
        row = cls._write_table(ws, row, "2.3.- Becarios y personal en formacion", ["Nro.", "Apellido y nombre", "Tipo de formacion", "Becas asociadas", "Fuente de financiamiento", "Carga horaria semanal"], becarios_rows, merge_span=8)
        equipamiento_rows = [[idx, item.denominacion, item.descripcion_breve, item.fecha_incorporacion, cls._money(item.monto_invertido)] for idx, item in enumerate(equipamiento, start=1)]
        row = cls._write_table(ws, row, "3.- EQUIPAMIENTO DEL GRUPO", ["Nro.", "Denominacion del equipamiento", "Descripcion breve", "Fecha de incorporacion", "Monto invertido"], equipamiento_rows, accent=True, merge_span=8, date_cols={4}, money_cols={5})
        docs_rows = [[idx, doc.titulo, ", ".join(a.nombre_apellido for a in doc.autores if not hasattr(a, "deleted_at") or a.deleted_at is None) or "-", doc.editorial, doc.anio] for idx, doc in enumerate(documentacion, start=1)]
        row = cls._write_table(ws, row, "4.- DOCUMENTACION Y BIBLIOTECA", ["Nro.", "Titulo de la obra", "Autores", "Editorial", "Anio de publicacion"], docs_rows, merge_span=8)

        row = cls._write_section(ws, row, "II.- ACTIVIDADES DE I+D+I")
        proyectos_rows = []
        total_proyectos = 0.0
        for idx, proyecto in enumerate(proyectos, start=1):
            monto = cls._money(proyecto.monto_destinado)
            total_proyectos += monto
            investigadores_activos = [rel.investigador for rel in proyecto.participaciones_investigador if rel.deleted_at is None and rel.investigador is not None and rel.investigador.deleted_at is None]
            becarios_activos = [rel.becario for rel in proyecto.participaciones_becario if rel.deleted_at is None and rel.becario is not None and rel.becario.deleted_at is None]
            distinciones_texto = "; ".join(dist.descripcion for dist in proyecto.distinciones if dist.deleted_at is None) or "-"
            proyectos_rows.append([idx, proyecto.codigo_proyecto, proyecto.nombre_proyecto, proyecto.tipo_proyecto.nombre if proyecto.tipo_proyecto else "-", proyecto.fuente_financiamiento.nombre if proyecto.fuente_financiamiento else "-", monto, proyecto.fecha_inicio, proyecto.fecha_fin, ", ".join(inv.nombre_apellido for inv in investigadores_activos) or "-", ", ".join(bec.nombre_apellido for bec in becarios_activos) or "-", distinciones_texto])
        row = cls._write_table(ws, row, "5.- PROYECTOS DE INVESTIGACION", ["Nro.", "Codigo del proyecto", "Denominacion del proyecto", "Tipo de proyecto", "Fuente de financiamiento", "Monto destinado", "Fecha de inicio", "Fecha de finalizacion", "Investigadores vinculados", "Becarios vinculados", "Distinciones asociadas"], proyectos_rows, accent=True, merge_span=12, date_cols={7, 8}, money_cols={6})
        row = cls._write_totals(ws, row, "Total monto proyectos", [(6, total_proyectos)])
        participaciones_rows = [[idx, participacion.nombre_evento, participacion.forma_participacion, participacion.fecha, participacion.investigador.nombre_apellido if participacion.investigador else "-"] for idx, participacion in enumerate(participaciones, start=1)]
        row = cls._write_subsection(ws, row, "6.- RECONOCIMIENTOS, PARTICIPACIONES Y VISITAS ACADEMICAS", span=10)
        distinciones_rows = [[idx, distincion.fecha, distincion.descripcion, distincion.proyecto_investigacion.nombre_proyecto if distincion.proyecto_investigacion else "-"] for idx, distincion in enumerate(distinciones, start=1)]
        row = cls._write_table(ws, row, "6.1.- Distinciones y reconocimientos recibidos", ["Nro.", "Fecha", "Descripcion de la distincion", "Proyecto asociado"], distinciones_rows, merge_span=8, date_cols={2})
        row = cls._write_table(ws, row, "6.2.- Participaciones institucionales y academicas relevantes", ["Nro.", "Evento o actividad", "Forma de participacion", "Fecha", "Investigador participante"], participaciones_rows, merge_span=8, date_cols={4})
        visitas_rows = [[idx, visita.razon, visita.procedencia, visita.tipo_visita.nombre if visita.tipo_visita else "-", visita.fecha] for idx, visita in enumerate(visitas, start=1)]
        row = cls._write_table(ws, row, "6.3.- Visitantes del pais y del extranjero", ["Nro.", "Motivo o razon de la visita", "Procedencia", "Tipo de visita", "Fecha"], visitas_rows, merge_span=8, date_cols={5})
        reuniones_grouped = {}
        for trabajo in trabajos_reunion:
            tipo = trabajo.tipo_reunion_cientifica.nombre if trabajo.tipo_reunion_cientifica else "Sin tipo definido"
            reuniones_grouped.setdefault(tipo, []).append([len(reuniones_grouped.get(tipo, [])) + 1, trabajo.titulo_trabajo, trabajo.nombre_reunion, trabajo.procedencia, trabajo.fecha_inicio, cls._join_names(trabajo.investigadores)])
        row = cls._write_grouped_tables(ws, row, "7", "TRABAJOS PRESENTADOS EN CONGRESOS Y REUNIONES CIENTIFICAS CON REFERATO", list(reuniones_grouped.items()), ["Nro.", "Titulo del trabajo", "Reunion cientifica", "Institucion de procedencia", "Fecha de presentacion", "Investigadores participantes"], merge_span=10, date_cols={5})
        articulos_rows = [[idx, articulo.titulo, articulo.descripcion, articulo.fecha_publicacion] for idx, articulo in enumerate(articulos, start=1)]
        row = cls._write_table(ws, row, "8.- TRABAJOS REALIZADOS Y PUBLICADOS", ["Nro.", "Titulo del articulo", "Descripcion o sintesis", "Fecha de publicacion"], articulos_rows, merge_span=8, date_cols={4})
        revistas_rows = [[idx, trabajo.titulo_trabajo, trabajo.nombre_revista, trabajo.editorial, trabajo.issn, trabajo.pais, trabajo.tipo_reunion.nombre if trabajo.tipo_reunion else "-", trabajo.fecha, cls._join_names(trabajo.investigadores)] for idx, trabajo in enumerate(trabajos_revista, start=1)]
        row = cls._write_table(ws, row, "8.1.- Trabajos en revistas con referato", ["Nro.", "Titulo del trabajo", "Revista", "Editorial", "ISSN", "Pais", "Tipo de publicacion", "Fecha", "Investigadores participantes"], revistas_rows, merge_span=10, date_cols={8})
        registros_grouped = {}
        for registro in registros:
            tipo = registro.tipo_registro.nombre if registro.tipo_registro else "Sin tipo definido"
            registros_grouped.setdefault(tipo, []).append([len(registros_grouped.get(tipo, [])) + 1, registro.nombre_articulo, registro.organismo_registrante, registro.fecha_registro])
        row = cls._write_grouped_tables(ws, row, "9", "REGISTROS Y PATENTES", list(registros_grouped.items()), ["Nro.", "Nombre o titulo registrado", "Organismo registrante", "Fecha de registro"], merge_span=8, date_cols={4})

        row = cls._write_section(ws, row, "III.- ACTIVIDADES EN DOCENCIA")
        actividades_rows = [[idx, actividad.curso, actividad.institucion, actividad.investigador.nombre_apellido if actividad.investigador else "-", cls._active_grado_nombre(actividad), actividad.rol_actividad.nombre if actividad.rol_actividad else "-", actividad.fecha_inicio, actividad.fecha_fin] for idx, actividad in enumerate(actividades, start=1)]
        row = cls._write_table(ws, row, "10.- ACTIVIDADES EN DOCENCIA", ["Nro.", "Curso o actividad", "Institucion", "Investigador responsable", "Grado academico vigente", "Rol desempenado", "Fecha de inicio", "Fecha de finalizacion"], actividades_rows, merge_span=10, date_cols={7, 8})

        row = cls._write_section(ws, row, "IV.- VINCULACION CON EL MEDIO SOCIO PRODUCTIVO")
        transferencias_grouped = {}
        total_transferencias = 0.0
        for transferencia in transferencias:
            monto = cls._money(transferencia.monto)
            total_transferencias += monto
            tipo = transferencia.tipo_contrato_transferencia.nombre if transferencia.tipo_contrato_transferencia else "Sin tipo definido"
            adoptantes = ", ".join(sorted({participacion.adoptante.nombre for participacion in transferencia.participaciones if participacion.deleted_at is None and participacion.adoptante is not None})) or "-"
            transferencias_grouped.setdefault(tipo, []).append([len(transferencias_grouped.get(tipo, [])) + 1, transferencia.numero_transferencia, transferencia.denominacion, transferencia.demandante, monto, transferencia.fecha_inicio, transferencia.fecha_fin, adoptantes])
        row = cls._write_grouped_tables(ws, row, "11", "TRANSFERENCIAS SOCIO-PRODUCTIVAS", list(transferencias_grouped.items()), ["Nro.", "Numero de transferencia", "Denominacion", "Demandante", "Monto comprometido", "Fecha de inicio", "Fecha de finalizacion", "Adoptantes vinculados"], accent=True, merge_span=10, date_cols={6, 7}, money_cols={5})
        row = cls._write_totals(ws, row, "Total monto transferencias", [(5, total_transferencias)])

        row = cls._write_section(ws, row, "V.- INFORME SOBRE RENDICION DE CUENTAS")
        erogaciones_grouped = {}
        total_ingresos = 0.0
        total_egresos = 0.0
        for erogacion in erogaciones:
            ingresos = cls._money(erogacion.ingresos)
            egresos = cls._money(erogacion.egresos)
            total_ingresos += ingresos
            total_egresos += egresos
            saldo = ingresos - egresos
            tipo = erogacion.tipo_erogacion.nombre if erogacion.tipo_erogacion else "Sin tipo definido"
            erogaciones_grouped.setdefault(tipo, []).append([len(erogaciones_grouped.get(tipo, [])) + 1, erogacion.numero_erogacion, erogacion.fecha, erogacion.fuente_financiamiento.nombre if erogacion.fuente_financiamiento else "-", ingresos, egresos, saldo])
        row = cls._write_grouped_tables(ws, row, "12", "RESUMEN DE INGRESOS Y EGRESOS (EROGACIONES)", list(erogaciones_grouped.items()), ["Nro.", "Numero de erogacion", "Fecha", "Fuente de financiamiento", "Ingresos", "Egresos", "Saldo resultante"], accent=True, merge_span=9, date_cols={3}, money_cols={5, 6, 7})
        row = cls._write_totals(ws, row, "Totales erogaciones", [(5, total_ingresos), (6, total_egresos), (7, total_ingresos - total_egresos)])

        row = cls._write_section(ws, row, "VI.- PROGRAMA DE ACTIVIDADES FUTURAS")
        planificaciones_rows = [[idx, plan.anio, plan.descripcion] for idx, plan in enumerate(planificaciones, start=1)]
        row = cls._write_table(ws, row, "13.- PLANIFICACIONES FUTURAS DEL GRUPO", ["Nro.", "Anio", "Descripcion"], planificaciones_rows, merge_span=8)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
